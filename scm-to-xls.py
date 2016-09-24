#!/usr/bin/env python3

from os import path, getcwd
from sys import exit

from pygit2 import Repository
from pygit2 import GIT_SORT_TIME, GIT_DIFF_STATS_FULL

import hglib

from openpyxl import Workbook
from openpyxl.styles import colors, Font, Color, Alignment, Border, Side, PatternFill
from openpyxl import cell

from optparse import OptionParser
from datetime import datetime

from string import ascii_uppercase


USAGE="""
This script allows to export version history from various Source Control Management
tools to excel spreadsheets (.xlsx). The main purpose is to allow a quick an easy
export for impact analysis."""


class LogEntry:
    def __init__(self):
        self.id = None
        self.msg = None
        self.author = None
        self.email = None
        self.time = None
        self.diff = None


class ScmAccessor:
    def __init__(self, repo_path, start_rev=None):
        self._repo_path = repo_path
        self._scm = None
        self._start_rev  = start_rev

    def get_log(self):
        raise RuntimeError("Not meant to be called in parent class")


class GitAccessor(ScmAccessor):
    def __init__(self, repo_path, start_rev=None):
        super().__init__(repo_path=repo_path, start_rev=start_rev)
        self._scm = Repository(path.join(repo_path))

    def get_log(self):
        data = []

        for c in self._scm.walk(self._scm.head.target, GIT_SORT_TIME):
            diff = self._scm.diff(c, c.parents[0]).stats.format(GIT_DIFF_STATS_FULL, 1) if c.parents else ""

            diff = diff.splitlines()
            if len(diff) >= 1:
                diff = diff[:-1]

            stripped_diff = [ d.split("|")[0].strip() for d in diff ]

            e = LogEntry()
            e.id = c.id
            e.msg = c.message.strip("\n")
            e.author = c.committer.name
            e.email = c.committer.email
            e.time = datetime.fromtimestamp(c.commit_time)
            e.diff = stripped_diff
            data.append(e)

            if self._start_rev and c.id.hex == self._start_rev:
                break

        return data


class HgAccessor(ScmAccessor):
    def __init__(self, repo_path, start_rev=None):
        super().__init__(repo_path=repo_path, start_rev=start_rev)
        self._scm = hglib.open(path.join(repo_path))

    def get_log(self):
        data = []

        for c in self._scm.log():
            #diff = self._scm.diff(c, c.parents[0]).stats.format(GIT_DIFF_STATS_FULL, 1) if c.parents else ""

            #diff = diff.splitlines()
            #if len(diff) >= 1:
            #    diff = diff[:-1]

            #stripped_diff = [ d.split("|")[0].strip() for d in diff ]

            tag = "" if not c.tags else " - " + str(c.tags, 'utf-8')

            e = LogEntry()
            e.id = str(c.rev, 'utf-8') + tag + " - " + str(c.node, 'utf-8')
            e.msg = str(c.desc, 'utf-8')
            e.author = str(c.author, 'utf-8').split("<")[0]
            e.email = str(c.author, 'utf-8').split("<")[1].rstrip(">")
            e.time = c.date
            e.diff = []
            data.append(e)

            if self._start_rev and c.id.hex == self._start_rev:
                break

        return data


class SvnAccessor(ScmAccessor):
    def __init__(self, repo_path, start_rev):
        super().__init__(repo_path=repo_path, start_rev=start_rev)
        raise NotImplementedError("Implement me!")



class Writer:
    def __init__(self, filename):
        self._filename = filename
        self._workbook = Workbook()
        self._columns = ["Date", "Commit-ID", "Message", "Author", "Changed files"]

        ws = self._workbook.active
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20

    def write_header(self):
        ws = self._workbook.active

        ws.append([])
        ws.append(self._columns)

        for c in ascii_uppercase[:len(self._columns)]:
            ws[c+str(ws.max_row)].font = Font(bold=True)
            ws[c+str(ws.max_row)].fill = PatternFill("solid", fgColor="000099ff")
            ws[c+str(ws.max_row)].border = Border(top=Side(border_style="thin", color="00000000"),
                                                  left=Side(border_style="thin", color="00000000"),
                                                  right=Side(border_style="thin", color="00000000"),
                                                  bottom=Side(border_style="thin", color="00000000"))

    def write_data(self):
        raise RuntimeError("Not to be called in base class")

    def write_data(self, accessor):
        ws = self._workbook.active

        for d in accessor.get_log():
            ws.append([d.time.strftime("%Y-%m-%d %H:%M:%S"), str(d.id), str(d.msg), str(d.author), "\n".join(d.diff)])

            for c in ascii_uppercase[:len(self._columns)]:
                ws[c+str(ws.max_row)].alignment = Alignment(vertical="top")

            ws["B"+str(ws.max_row)].font = Font(name="Monospace", size=8)
            ws["C"+str(ws.max_row)].font = Font(size=10)
            ws["C"+str(ws.max_row)].alignment = Alignment(wrap_text=True, shrink_to_fit=True, vertical="top")
            #ws.row_dimensions[ws.max_row].height = d.msg.count("\n") * 15

            for c in ascii_uppercase[:len(self._columns)]:
                ws[c+str(ws.max_row)].border = Border(left=Side(border_style="thin", color="00000000"),
                                                      right=Side(border_style="thin", color="00000000"))

        for c in ascii_uppercase[:len(self._columns)]:
            ws[c+str(ws.max_row)].border = Border(left=Side(border_style="thin", color="00000000"),
                                                  right=Side(border_style="thin", color="00000000"),
                                                  bottom=Side(border_style="thin", color="00000000"))

    def save(self):
        self._workbook.save(self._filename)


class CommitHistoryWriter(Writer):
    def __init__(self, filename):
        super().__init__(filename)

    def write_header(self):
        ws = self._workbook.active
        ws['A1'] = "Commit history"
        ws['B1'] = "Generated on %s" % datetime.now().isoformat(" ")

        super().write_header()


class ImpactStatementWriter(Writer):
    def __init__(self, filename):
        super().__init__(filename=filename)

    def write_header(self):
        ws = self._workbook.active

        ws['A1'] = "Impact statement"
        ws['B1'] = "Generated on %s" % datetime.now().isoformat(" ")
        self._columns.extend(["Affected testcases", "Tested with version"])
        super().write_header()



def main():
    parser = OptionParser(usage=USAGE)

    parser.add_option("-o", "--outfile", dest="outfile",
                      help="Output file to write to.")
    parser.add_option("-s", "--scm", dest="scm",
                      help="SCM to use. Supported options: git, hg, svn",
                      default="git")
    parser.add_option("-I", "--impact", dest="impact",
                      help="If set, an impact statement is generated.",
                      action="store_true")
    parser.add_option("-H", "--history", dest="history",
                      help="If set, the commit history is exported.",
                      action="store_true")
    parser.add_option("-d", "--directory", dest="repo",
                      help="Path to the repository. If omitted, the location "
                            "the script is placed at is used.")
    parser.add_option("-r", "--rev", dest="rev",
                      help="Revision to start. If omitted, the complete history "
                           "is used. Dependent on the SCM used!", default=None)

    options, args = parser.parse_args()

    if not options.history and not options.impact:
        print("Not sure what to do. Impact analysis (-I) or commit history (-H)?")
        exit(-1)

    if not options.outfile:
        print("Output filename missing.")
        exit(-1)

    rpath = options.repo if options.repo else getcwd()
    rev = options.rev

    accessor = None

    if options.scm == "git":
        accessor = GitAccessor(rpath, rev)
    elif options.scm == "hg":
        accessor = HgAccessor(rpath, rev)
    elif options.scm == "svn":
        accessor = SvnAccessor(rpath, rev)

    outfile = options.outfile

    if not outfile.endswith(".xlsx"):
        outfile += ".xlsx"

    if options.history:
        print("Generating commit history...")
        writer = CommitHistoryWriter("History-" + outfile)
    elif options.impact:
        print("Generating impact statement...")
        writer = ImpactStatementWriter("Impacts-" + outfile)

    writer.write_header()
    writer.write_data(accessor=accessor)
    writer.save()

    print("...done.")

if __name__ == "__main__":
    main()
